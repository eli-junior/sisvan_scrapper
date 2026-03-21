"""
Script de validação e consolidação dos dados SISVAN.

Pastas suportadas:
  - dados/por_sexo      → consolida em dados/consolidado_por_sexo.xlsx
  - dados/por_raca_cor  → consolida em dados/consolidado_por_raca_cor.xlsx

Validações por arquivo:
  - Ano    : igual ao nome do arquivo
  - Mês    : TODOS
  - Sexo   : TODOS (raca_cor) | FEMININO/MASCULINO (por_sexo)
  - Raça/Cor: igual ao nome do arquivo (apenas raca_cor)
  - Abrangência: BRASIL
  - Tipo   : Consumo de Alimentos Ultraprocessados
  - Faixa  : Total de Crianças de 5 a 9 anos

Comportamento incremental: na próxima execução, preenche apenas os dados
faltantes (arquivos que ainda não foram validados com sucesso).

Uso:
  python validate_sisvan.py [--sexo] [--raca] [--todos]
"""

import re
import sys
import logging
import argparse
import unicodedata
from pathlib import Path
from html.parser import HTMLParser

import openpyxl
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# ---------------------------------------------------------------------------
# Configuração geral
# ---------------------------------------------------------------------------

BASE_DIR = Path(__file__).parent / "dados"
LOG_FILE = Path(__file__).parent / "logs" / "validate_sisvan.log"

ANOS = list(range(2015, 2025))

SEXOS = ["FEMININO", "MASCULINO"]

# Nome canonical (usado no nome do arquivo) → label esperado no HTML
RACAS = {
    "Branca":   "BRANCA",
    "Preta":    "PRETA",
    "Amarela":  "AMARELA",
    "Parda":    "PARDA",
    "Indigena": "INDÍGENA",
}

# Strings esperadas nos cabeçalhos da tabela
EXPECT_MES          = "TODOS"
EXPECT_ABRANGENCIA  = "BRASIL"
EXPECT_TIPO         = "consumo de alimentos ultraprocessados"
EXPECT_FAIXA        = "crianças de 5 a 9 anos"

# ---------------------------------------------------------------------------
# Logging
# ---------------------------------------------------------------------------

LOG_FILE.parent.mkdir(parents=True, exist_ok=True)
logging.basicConfig(
    level=logging.INFO,
    format="%(asctime)s [%(levelname)s] %(message)s",
    datefmt="%H:%M:%S",
    handlers=[
        logging.StreamHandler(sys.stdout),
        logging.FileHandler(LOG_FILE, encoding="utf-8", mode="a"),
    ],
)
log = logging.getLogger(__name__)


# ---------------------------------------------------------------------------
# Parser HTML
# ---------------------------------------------------------------------------

class SisvanParser(HTMLParser):
    """Extrai metadados e dados da tabela do relatório HTML do SISVAN."""

    def __init__(self):
        super().__init__()

        # --- estado de parsing ---
        self._collecting_meta = False
        self._in_strong = False
        self._last_strong: str | None = None
        self._in_td_th = False
        self._in_thead = False
        self._in_tbody = False
        self._in_tr = False
        self._current_cell: list[str] = []
        self._current_row_cells: list[str] = []

        # --- resultado ---
        self.meta_ano: str | None = None
        self.meta_mes: str | None = None
        self.meta_sexo: str | None = None
        self.meta_raca_cor: str | None = None

        self.form_ano: str | None = None
        self.form_mes: str | None = None    # "99" = TODOS
        self.form_filtro: str | None = None  # "F" = BRASIL

        self.thead_texts: list[str] = []
        self.tbody_cells: list[str] = []    # células da linha BRASIL
        self._brasil_row_found = False

    # ---- helpers -----------------------------------------------------------

    @staticmethod
    def _clean(text: str) -> str:
        return re.sub(r"\s+", " ", text).strip()

    # ---- callbacks ---------------------------------------------------------

    def handle_starttag(self, tag, attrs):
        attr = dict(attrs)

        if tag == "div" and "box-body" in attr.get("class", ""):
            self._collecting_meta = True

        if tag == "strong" and self._collecting_meta:
            self._in_strong = True
            self._current_cell = []

        if tag == "thead":
            self._in_thead = True
        if tag == "tbody":
            self._in_tbody = True
        if tag in ("td", "th"):
            self._in_td_th = True
            self._current_cell = []
        if tag == "tr" and self._in_tbody:
            self._in_tr = True
            self._current_row_cells = []

        if tag == "input" and attr.get("type") == "hidden":
            name, value = attr.get("name", ""), attr.get("value", "")
            if name == "nuAno":
                self.form_ano = value
            elif name == "nuMes[]":
                self.form_mes = value
            elif name == "tpFiltro":
                self.form_filtro = value

    def handle_endtag(self, tag):
        if tag == "strong" and self._in_strong:
            self._in_strong = False
            text = self._clean("".join(self._current_cell))
            self._last_strong = text

        if tag in ("td", "th") and self._in_td_th:
            self._in_td_th = False
            text = self._clean("".join(self._current_cell))
            if self._in_thead:
                if text:
                    self.thead_texts.append(text)
            elif self._in_tbody and self._in_tr:
                self._current_row_cells.append(text)

        if tag == "tr" and self._in_tbody and self._in_tr:
            self._in_tr = False
            if "BRASIL" in " ".join(self._current_row_cells).upper() and not self._brasil_row_found:
                self._brasil_row_found = True
                self.tbody_cells = self._current_row_cells[:]

        if tag == "thead":
            self._in_thead = False
        if tag == "tbody":
            self._in_tbody = False

    def handle_data(self, data):
        if self._in_strong:
            self._current_cell.append(data)
        elif self._in_td_th:
            self._current_cell.append(data)
        elif self._collecting_meta and self._last_strong:
            text = data.strip()
            if not text:
                return
            label = self._last_strong.rstrip(":").strip().upper()
            if label == "ANO":
                m = re.search(r"(\d{4})", text)
                if m:
                    self.meta_ano = m.group(1)
            elif label == "MÊS":
                self.meta_mes = text.strip(" -")
            elif label == "SEXO":
                self.meta_sexo = text
            elif label == "RAÇA E COR":
                self.meta_raca_cor = text
            self._last_strong = None


# ---------------------------------------------------------------------------
# Leitura e validação
# ---------------------------------------------------------------------------

def parse_xls(filepath: Path) -> SisvanParser:
    """Lê o arquivo HTML (extensão .xls) e faz parse."""
    raw = filepath.read_bytes()
    for enc in ("utf-8", "latin-1", "cp1252"):
        try:
            content = raw.decode(enc)
            break
        except UnicodeDecodeError:
            continue
    else:
        content = raw.decode("utf-8", errors="replace")

    p = SisvanParser()
    p.feed(content)
    return p


def _extrair_dados(parser: SisvanParser) -> tuple[str | None, str | None, str | None]:
    """Retorna (total, percentual, acompanhados) da linha BRASIL."""
    numeric = [c for c in parser.tbody_cells if c.upper() != "BRASIL"]
    total = numeric[0] if len(numeric) > 0 else None
    pct   = numeric[1] if len(numeric) > 1 else None
    acomp = numeric[2] if len(numeric) > 2 else None
    return total, pct, acomp


def validar_por_sexo(filepath: Path, ano: int, sexo: str) -> dict:
    """Valida arquivo da pasta por_sexo."""
    divergencias = []
    p = parse_xls(filepath)

    # Ano
    ano_enc = p.meta_ano or p.form_ano
    if not ano_enc:
        divergencias.append("ANO: não encontrado")
    elif str(ano) != str(ano_enc).strip():
        divergencias.append(f"ANO: esperado {ano}, encontrado '{ano_enc}'")

    # Sexo
    if not p.meta_sexo:
        divergencias.append("SEXO: não encontrado")
    elif sexo.upper() != p.meta_sexo.upper():
        divergencias.append(f"SEXO: esperado '{sexo}', encontrado '{p.meta_sexo}'")

    # Mês
    mes_enc = p.meta_mes or ("TODOS" if p.form_mes == "99" else p.form_mes)
    if not mes_enc or EXPECT_MES.upper() not in str(mes_enc).upper():
        divergencias.append(f"MÊS: esperado 'TODOS', encontrado '{mes_enc}'")

    # BRASIL
    if not p._brasil_row_found and p.form_filtro != "F":
        divergencias.append("ABRANGÊNCIA: linha BRASIL não encontrada")

    # Tipo
    if not any(EXPECT_TIPO in t.lower() for t in p.thead_texts):
        divergencias.append("TIPO: 'Consumo de Alimentos Ultraprocessados' não encontrado")

    # Faixa etária
    if not any(EXPECT_FAIXA in t.lower() for t in p.thead_texts):
        divergencias.append("FAIXA ETÁRIA: 'Crianças de 5 a 9 anos' não encontrado")

    # Dados
    total, pct, acomp = _extrair_dados(p)
    if total is None:
        divergencias.append("DADOS: valores numéricos não encontrados na linha BRASIL")

    return {
        "ano": ano, "sexo": sexo,
        "total": total, "percentual": pct, "acompanhados": acomp,
        "divergencias": divergencias,
        "valido": len(divergencias) == 0,
    }


def _sem_acento(texto: str) -> str:
    """Remove acentos e normaliza para comparação."""
    return unicodedata.normalize("NFD", texto).encode("ascii", "ignore").decode("ascii").upper()


def validar_por_raca(filepath: Path, ano: int, raca_nome: str) -> dict:
    """Valida arquivo da pasta por_raca_cor.

    raca_nome: nome canonical do arquivo (ex: 'Amarela', 'Indigena')
    """
    divergencias = []
    p = parse_xls(filepath)

    # Ano
    ano_enc = p.meta_ano or p.form_ano
    if not ano_enc:
        divergencias.append("ANO: não encontrado")
    elif str(ano) != str(ano_enc).strip():
        divergencias.append(f"ANO: esperado {ano}, encontrado '{ano_enc}'")

    # Sexo = TODOS
    if p.meta_sexo and p.meta_sexo.upper() != "TODOS":
        divergencias.append(f"SEXO: esperado 'TODOS', encontrado '{p.meta_sexo}'")

    # Raça e Cor — comparação sem acento para lidar com "Indígena" vs "Indigena"
    label_esperado = RACAS.get(raca_nome, raca_nome.upper())
    if not p.meta_raca_cor:
        divergencias.append(f"RAÇA E COR: não encontrado (esperado '{label_esperado}')")
    elif _sem_acento(label_esperado) != _sem_acento(p.meta_raca_cor):
        divergencias.append(
            f"RAÇA E COR: esperado '{label_esperado}', encontrado '{p.meta_raca_cor}'"
        )

    # Mês
    mes_enc = p.meta_mes or ("TODOS" if p.form_mes == "99" else p.form_mes)
    if not mes_enc or EXPECT_MES.upper() not in str(mes_enc).upper():
        divergencias.append(f"MÊS: esperado 'TODOS', encontrado '{mes_enc}'")

    # BRASIL
    if not p._brasil_row_found and p.form_filtro != "F":
        divergencias.append("ABRANGÊNCIA: linha BRASIL não encontrada")

    # Tipo
    if not any(EXPECT_TIPO in t.lower() for t in p.thead_texts):
        divergencias.append("TIPO: 'Consumo de Alimentos Ultraprocessados' não encontrado")

    # Faixa etária
    if not any(EXPECT_FAIXA in t.lower() for t in p.thead_texts):
        divergencias.append("FAIXA ETÁRIA: 'Crianças de 5 a 9 anos' não encontrado")

    # Dados
    total, pct, acomp = _extrair_dados(p)
    if total is None:
        divergencias.append("DADOS: valores numéricos não encontrados na linha BRASIL")

    return {
        "ano": ano, "raca": raca_nome,
        "total": total, "percentual": pct, "acompanhados": acomp,
        "divergencias": divergencias,
        "valido": len(divergencias) == 0,
    }


# ---------------------------------------------------------------------------
# Consolidação Excel — helpers de estilo
# ---------------------------------------------------------------------------

def _thin_border():
    s = Side(style="thin", color="9E9E9E")
    return Border(left=s, right=s, top=s, bottom=s)


def _center():
    return Alignment(horizontal="center", vertical="center")


def _wrap():
    return Alignment(horizontal="center", vertical="center", wrap_text=True)


# ---------------------------------------------------------------------------
# Consolidação: por_sexo
# ---------------------------------------------------------------------------

OUTPUT_SEXO = BASE_DIR / "consolidado_por_sexo.xlsx"

# Paleta de cores por sexo
_FILL_SEXO = {
    "FEMININO":  PatternFill("solid", fgColor="FCE4EC"),
    "MASCULINO": PatternFill("solid", fgColor="E3F2FD"),
}
_FILL_INVALID = PatternFill("solid", fgColor="FFCDD2")
_FILL_MISSING = PatternFill("solid", fgColor="FFF9C4")


def _carregar_existentes_sexo(output_file: Path) -> dict:
    """Lê o consolidado de sexo e devolve {(ano, sexo): {...}}."""
    existentes = {}
    if not output_file.exists():
        return existentes
    try:
        wb = openpyxl.load_workbook(output_file)
        ws = wb.active
        for row in ws.iter_rows(min_row=3, values_only=True):
            if row[0] is None:
                continue
            ano, sexo, total, pct, acomp = row[0], row[1], row[2], row[3], row[4]
            if ano and sexo:
                existentes[(int(ano), str(sexo).upper())] = {
                    "total": total, "percentual": pct, "acompanhados": acomp,
                }
    except Exception as e:
        log.warning(f"Não foi possível ler consolidado existente (sexo): {e}")
    return existentes


def _criar_wb_sexo():
    wb = Workbook()
    ws = wb.active
    ws.title = "por_sexo"

    hf = Font(name="Calibri", bold=True, color="FFFFFF", size=11)
    hfill = PatternFill("solid", fgColor="2E7D32")
    border = _thin_border()

    ws.merge_cells("A1:E1")
    ws["A1"].value = (
        "SISVAN – Consumo de Alimentos Ultraprocessados | "
        "Crianças de 5 a 9 anos | BRASIL | por Sexo"
    )
    ws["A1"].font = Font(name="Calibri", bold=True, color="FFFFFF", size=12)
    ws["A1"].fill = PatternFill("solid", fgColor="1B5E20")
    ws["A1"].alignment = _center()
    ws.row_dimensions[1].height = 22

    headers = ["Ano", "Sexo", "Total (Ultraprocessados)", "% (Ultraprocessados)", "Total Acompanhados(as)"]
    for col, h in enumerate(headers, 1):
        c = ws.cell(row=2, column=col, value=h)
        c.font = hf; c.fill = hfill; c.alignment = _wrap(); c.border = border
    ws.row_dimensions[2].height = 30

    for col, w in zip("ABCDE", [8, 14, 24, 20, 24]):
        ws.column_dimensions[col].width = w

    return wb, ws


def _escrever_linha_sexo(ws, row: int, ano, sexo, total, pct, acomp, invalido=False, faltando=False):
    border = _thin_border()
    if faltando:
        fill = _FILL_MISSING
    elif invalido:
        fill = _FILL_INVALID
    else:
        fill = _FILL_SEXO.get(str(sexo).upper(), PatternFill())

    for col, val in enumerate([ano, sexo, total, pct, acomp], 1):
        c = ws.cell(row=row, column=col, value=val)
        c.alignment = _center(); c.border = border; c.fill = fill


def processar_por_sexo():
    log.info("")
    log.info("=" * 60)
    log.info("PARTE 1: POR SEXO")
    log.info("=" * 60)

    source_dir = BASE_DIR / "por_sexo"
    if not source_dir.exists():
        log.error(f"Pasta não encontrada: {source_dir}")
        return

    existentes = _carregar_existentes_sexo(OUTPUT_SEXO)
    log.info(f"Consolidado existente: {len(existentes)} registros")

    wb, ws = _criar_wb_sexo()
    total_arq = len(ANOS) * len(SEXOS)
    cont = {"ok": 0, "div": 0, "miss": 0, "skip": 0}
    divergencias_report = []
    row_num = 3

    for ano in ANOS:
        for sexo in SEXOS:
            idx = (ano - ANOS[0]) * len(SEXOS) + SEXOS.index(sexo) + 1
            chave = (ano, sexo.upper())
            nome = f"sexo_{ano}_{sexo}.xls"
            fp = source_dir / nome

            if not fp.exists():
                log.warning(f"  [{idx}/{total_arq}] FALTANDO: {nome}")
                cont["miss"] += 1
                _escrever_linha_sexo(ws, row_num, ano, sexo, "ARQUIVO FALTANDO", None, None, faltando=True)
                row_num += 1
                continue

            # Reutilizar dados válidos do consolidado anterior
            if chave in existentes:
                d = existentes[chave]
                if d["total"] not in (None, "INVÁLIDO", "ARQUIVO FALTANDO"):
                    log.info(f"  [{idx}/{total_arq}] JÁ PROCESSADO: {nome}")
                    cont["skip"] += 1
                    _escrever_linha_sexo(ws, row_num, ano, sexo, d["total"], d["percentual"], d["acompanhados"])
                    row_num += 1
                    continue

            log.info(f"  [{idx}/{total_arq}] Validando: {nome}")
            r = validar_por_sexo(fp, ano, sexo)

            if r["valido"]:
                cont["ok"] += 1
                log.info(f"    ✓ OK | Total={r['total']} | %={r['percentual']} | Acomp={r['acompanhados']}")
                _escrever_linha_sexo(ws, row_num, ano, sexo, r["total"], r["percentual"], r["acompanhados"])
            else:
                cont["div"] += 1
                log.warning(f"    ✗ DIVERGÊNCIAS em {nome}:")
                for d in r["divergencias"]:
                    log.warning(f"      - {d}")
                divergencias_report.append((nome, r["divergencias"]))
                _escrever_linha_sexo(ws, row_num, ano, sexo, "INVÁLIDO", None, None, invalido=True)

            row_num += 1

    OUTPUT_SEXO.parent.mkdir(parents=True, exist_ok=True)
    wb.save(OUTPUT_SEXO)

    log.info(f"\n  Resultado: {cont['ok']} validados | {cont['skip']} reutilizados | "
             f"{cont['div']} divergências | {cont['miss']} faltando")
    log.info(f"  Salvo em: {OUTPUT_SEXO.resolve()}")

    if divergencias_report:
        log.info("\n  ARQUIVOS COM DIVERGÊNCIAS (sexo):")
        for nome, divs in divergencias_report:
            log.warning(f"    {nome}:")
            for d in divs:
                log.warning(f"      - {d}")

    return cont


# ---------------------------------------------------------------------------
# Consolidação: por_raca_cor
# ---------------------------------------------------------------------------

OUTPUT_RACA = BASE_DIR / "consolidado_por_raca_cor.xlsx"

# Uma cor por raça/cor
_FILL_RACA = {
    "Branca":   PatternFill("solid", fgColor="F5F5F5"),
    "Preta":    PatternFill("solid", fgColor="D7CCC8"),
    "Amarela":  PatternFill("solid", fgColor="FFF9C4"),
    "Parda":    PatternFill("solid", fgColor="FFCCBC"),
    "Indigena": PatternFill("solid", fgColor="DCEDC8"),
}


def _carregar_existentes_raca(output_file: Path) -> dict:
    """Lê o consolidado de raca e devolve {(ano, raca): {...}}."""
    existentes = {}
    if not output_file.exists():
        return existentes
    try:
        wb = openpyxl.load_workbook(output_file)
        ws = wb.active
        for row in ws.iter_rows(min_row=3, values_only=True):
            if row[0] is None:
                continue
            ano, raca, total, pct, acomp = row[0], row[1], row[2], row[3], row[4]
            if ano and raca:
                existentes[(int(ano), str(raca))] = {
                    "total": total, "percentual": pct, "acompanhados": acomp,
                }
    except Exception as e:
        log.warning(f"Não foi possível ler consolidado existente (raca): {e}")
    return existentes


def _criar_wb_raca():
    wb = Workbook()
    ws = wb.active
    ws.title = "por_raca_cor"

    hf = Font(name="Calibri", bold=True, color="FFFFFF", size=11)
    hfill = PatternFill("solid", fgColor="4527A0")  # roxo escuro
    border = _thin_border()

    ws.merge_cells("A1:E1")
    ws["A1"].value = (
        "SISVAN – Consumo de Alimentos Ultraprocessados | "
        "Crianças de 5 a 9 anos | BRASIL | por Raça/Cor"
    )
    ws["A1"].font = Font(name="Calibri", bold=True, color="FFFFFF", size=12)
    ws["A1"].fill = PatternFill("solid", fgColor="311B92")
    ws["A1"].alignment = _center()
    ws.row_dimensions[1].height = 22

    headers = ["Ano", "Raça/Cor", "Total (Ultraprocessados)", "% (Ultraprocessados)", "Total Acompanhados(as)"]
    for col, h in enumerate(headers, 1):
        c = ws.cell(row=2, column=col, value=h)
        c.font = hf; c.fill = hfill; c.alignment = _wrap(); c.border = border
    ws.row_dimensions[2].height = 30

    for col, w in zip("ABCDE", [8, 14, 24, 20, 24]):
        ws.column_dimensions[col].width = w

    return wb, ws


def _escrever_linha_raca(ws, row: int, ano, raca, total, pct, acomp, invalido=False, faltando=False):
    border = _thin_border()
    if faltando:
        fill = _FILL_MISSING
    elif invalido:
        fill = _FILL_INVALID
    else:
        fill = _FILL_RACA.get(str(raca), PatternFill())

    for col, val in enumerate([ano, raca, total, pct, acomp], 1):
        c = ws.cell(row=row, column=col, value=val)
        c.alignment = _center(); c.border = border; c.fill = fill


def processar_por_raca():
    log.info("")
    log.info("=" * 60)
    log.info("PARTE 2: POR RAÇA/COR")
    log.info("=" * 60)

    source_dir = BASE_DIR / "por_raca_cor"
    if not source_dir.exists():
        log.error(f"Pasta não encontrada: {source_dir}")
        return

    existentes = _carregar_existentes_raca(OUTPUT_RACA)
    log.info(f"Consolidado existente: {len(existentes)} registros")

    wb, ws = _criar_wb_raca()
    racas_lista = list(RACAS.keys())
    total_arq = len(ANOS) * len(racas_lista)
    cont = {"ok": 0, "div": 0, "miss": 0, "skip": 0}
    divergencias_report = []
    row_num = 3

    for ano in ANOS:
        for raca_nome in racas_lista:
            idx = (ano - ANOS[0]) * len(racas_lista) + racas_lista.index(raca_nome) + 1
            chave = (ano, raca_nome)
            nome = f"raca_{ano}_{raca_nome}.xls"
            fp = source_dir / nome

            if not fp.exists():
                log.warning(f"  [{idx}/{total_arq}] FALTANDO: {nome}")
                cont["miss"] += 1
                _escrever_linha_raca(ws, row_num, ano, raca_nome, "ARQUIVO FALTANDO", None, None, faltando=True)
                row_num += 1
                continue

            # Reutilizar dados válidos do consolidado anterior
            if chave in existentes:
                d = existentes[chave]
                if d["total"] not in (None, "INVÁLIDO", "ARQUIVO FALTANDO"):
                    log.info(f"  [{idx}/{total_arq}] JÁ PROCESSADO: {nome}")
                    cont["skip"] += 1
                    _escrever_linha_raca(ws, row_num, ano, raca_nome, d["total"], d["percentual"], d["acompanhados"])
                    row_num += 1
                    continue

            log.info(f"  [{idx}/{total_arq}] Validando: {nome}")
            r = validar_por_raca(fp, ano, raca_nome)

            if r["valido"]:
                cont["ok"] += 1
                log.info(f"    ✓ OK | Total={r['total']} | %={r['percentual']} | Acomp={r['acompanhados']}")
                _escrever_linha_raca(ws, row_num, ano, raca_nome, r["total"], r["percentual"], r["acompanhados"])
            else:
                cont["div"] += 1
                log.warning(f"    ✗ DIVERGÊNCIAS em {nome}:")
                for d in r["divergencias"]:
                    log.warning(f"      - {d}")
                divergencias_report.append((nome, r["divergencias"]))
                _escrever_linha_raca(ws, row_num, ano, raca_nome, "INVÁLIDO", None, None, invalido=True)

            row_num += 1

    OUTPUT_RACA.parent.mkdir(parents=True, exist_ok=True)
    wb.save(OUTPUT_RACA)

    log.info(f"\n  Resultado: {cont['ok']} validados | {cont['skip']} reutilizados | "
             f"{cont['div']} divergências | {cont['miss']} faltando")
    log.info(f"  Salvo em: {OUTPUT_RACA.resolve()}")

    if divergencias_report:
        log.info("\n  ARQUIVOS COM DIVERGÊNCIAS (raca_cor):")
        for nome, divs in divergencias_report:
            log.warning(f"    {nome}:")
            for d in divs:
                log.warning(f"      - {d}")

    return cont


# ---------------------------------------------------------------------------
# Parser específico para por_regiao (múltiplas linhas no tbody)
# ---------------------------------------------------------------------------

REGIOES = ["CENTRO-OESTE", "NORDESTE", "NORTE", "SUDESTE", "SUL"]


class SisvanParserRegiao(SisvanParser):
    """Estende o parser base para capturar TODAS as linhas do tbody (uma por região)."""

    def __init__(self):
        super().__init__()
        # Substitui tbody_cells por dict {regiao: (total, pct, acomp)}
        self.regioes_encontradas: dict[str, tuple] = {}
        self._all_rows: list[list[str]] = []  # todas as linhas brutas do tbody

    def handle_endtag(self, tag):
        # Captura de tr: registra TODA linha, não apenas a de BRASIL
        if tag in ("td", "th") and self._in_td_th:
            self._in_td_th = False
            text = self._clean("".join(self._current_cell))
            if self._in_thead:
                if text:
                    self.thead_texts.append(text)
            elif self._in_tbody and self._in_tr:
                self._current_row_cells.append(text)

        if tag == "tr" and self._in_tbody and self._in_tr:
            self._in_tr = False
            row = self._current_row_cells[:]
            self._all_rows.append(row)
            row_text = " ".join(row).upper()
            # Identificar a região desta linha
            for regiao in REGIOES:
                if regiao in row_text:
                    numeric = [c for c in row if c.upper() not in (regiao, "TOTAL BRASIL", "BRASIL")
                               and not c.upper().startswith("TOTAL")]
                    # Pegar apenas as células que não são o nome da região
                    vals = [c for c in row if c not in ("", ) and c.upper() not in row_text.split()[:2]]
                    # Simples: pegar todas as células que não contêm o nome da região
                    nums = [c for c in row if c.strip() and c.upper() not in (regiao,)
                            and not any(r in c.upper() for r in REGIOES + ["BRASIL"])]
                    if len(nums) >= 3:
                        self.regioes_encontradas[regiao] = (nums[0], nums[1], nums[2])
                    elif len(nums) == 2:
                        self.regioes_encontradas[regiao] = (nums[0], nums[1], None)
                    break

        if tag == "strong" and self._in_strong:
            self._in_strong = False
            text = self._clean("".join(self._current_cell))
            self._last_strong = text

        if tag == "thead":
            self._in_thead = False
        if tag == "tbody":
            self._in_tbody = False


def parse_xls_regiao(filepath: Path) -> SisvanParserRegiao:
    """Lê o arquivo HTML de região e faz parse."""
    raw = filepath.read_bytes()
    for enc in ("utf-8", "latin-1", "cp1252"):
        try:
            content = raw.decode(enc)
            break
        except UnicodeDecodeError:
            continue
    else:
        content = raw.decode("utf-8", errors="replace")

    p = SisvanParserRegiao()
    p.feed(content)
    return p


def validar_por_regiao(filepath: Path, ano: int) -> dict:
    """
    Valida arquivo da pasta por_regiao.
    Retorna dict com chaves: ano, regioes (dict), divergencias, valido.
    regioes = { 'CENTRO-OESTE': (total, pct, acomp), ... }
    """
    divergencias = []
    p = parse_xls_regiao(filepath)

    # Ano
    ano_enc = p.meta_ano or p.form_ano
    if not ano_enc:
        divergencias.append("ANO: não encontrado")
    elif str(ano) != str(ano_enc).strip():
        divergencias.append(f"ANO: esperado {ano}, encontrado '{ano_enc}'")

    # Mês = TODOS
    mes_enc = p.meta_mes or ("TODOS" if p.form_mes == "99" else p.form_mes)
    if not mes_enc or EXPECT_MES.upper() not in str(mes_enc).upper():
        divergencias.append(f"MÊS: esperado 'TODOS', encontrado '{mes_enc}'")

    # Sexo = TODOS
    if p.meta_sexo and p.meta_sexo.upper() != "TODOS":
        divergencias.append(f"SEXO: esperado 'TODOS', encontrado '{p.meta_sexo}'")

    # Agrupamento = REGIÃO (tpFiltro=R)
    if p.form_filtro and p.form_filtro != "R":
        divergencias.append(f"AGRUPAMENTO: esperado 'R' (REGIÃO), encontrado '{p.form_filtro}'")

    # Tipo
    if not any(EXPECT_TIPO in t.lower() for t in p.thead_texts):
        divergencias.append("TIPO: 'Consumo de Alimentos Ultraprocessados' não encontrado")

    # Faixa etária
    if not any(EXPECT_FAIXA in t.lower() for t in p.thead_texts):
        divergencias.append("FAIXA ETÁRIA: 'Crianças de 5 a 9 anos' não encontrado")

    # Verificar que todas as 5 regiões foram encontradas
    regioes_faltando = [r for r in REGIOES if r not in p.regioes_encontradas]
    if regioes_faltando:
        divergencias.append(f"REGIÕES FALTANDO: {', '.join(regioes_faltando)}")

    return {
        "ano": ano,
        "regioes": p.regioes_encontradas,
        "divergencias": divergencias,
        "valido": len(divergencias) == 0,
    }


# ---------------------------------------------------------------------------
# Consolidação: por_regiao
# ---------------------------------------------------------------------------

OUTPUT_REGIAO = BASE_DIR / "consolidado_por_regiao.xlsx"

_FILL_REGIAO = {
    "CENTRO-OESTE": PatternFill("solid", fgColor="E8F5E9"),
    "NORDESTE":     PatternFill("solid", fgColor="E3F2FD"),
    "NORTE":        PatternFill("solid", fgColor="FFF8E1"),
    "SUDESTE":      PatternFill("solid", fgColor="FCE4EC"),
    "SUL":          PatternFill("solid", fgColor="F3E5F5"),
}


def _carregar_existentes_regiao(output_file: Path) -> dict:
    """Lê o consolidado de região e devolve {(ano, regiao): {...}}."""
    existentes = {}
    if not output_file.exists():
        return existentes
    try:
        wb = openpyxl.load_workbook(output_file)
        ws = wb.active
        for row in ws.iter_rows(min_row=3, values_only=True):
            if row[0] is None:
                continue
            ano, regiao, total, pct, acomp = row[0], row[1], row[2], row[3], row[4]
            if ano and regiao:
                existentes[(int(ano), str(regiao).upper())] = {
                    "total": total, "percentual": pct, "acompanhados": acomp,
                }
    except Exception as e:
        log.warning(f"Não foi possível ler consolidado existente (regiao): {e}")
    return existentes


def _criar_wb_regiao():
    wb = Workbook()
    ws = wb.active
    ws.title = "por_regiao"

    hf = Font(name="Calibri", bold=True, color="FFFFFF", size=11)
    hfill = PatternFill("solid", fgColor="E65100")  # laranja escuro
    border = _thin_border()

    ws.merge_cells("A1:E1")
    ws["A1"].value = (
        "SISVAN – Consumo de Alimentos Ultraprocessados | "
        "Crianças de 5 a 9 anos | por Região"
    )
    ws["A1"].font = Font(name="Calibri", bold=True, color="FFFFFF", size=12)
    ws["A1"].fill = PatternFill("solid", fgColor="BF360C")
    ws["A1"].alignment = _center()
    ws.row_dimensions[1].height = 22

    headers = ["Ano", "Região", "Total (Ultraprocessados)", "% (Ultraprocessados)", "Total Acompanhados(as)"]
    for col, h in enumerate(headers, 1):
        c = ws.cell(row=2, column=col, value=h)
        c.font = hf; c.fill = hfill; c.alignment = _wrap(); c.border = border
    ws.row_dimensions[2].height = 30

    for col, w in zip("ABCDE", [8, 16, 24, 20, 24]):
        ws.column_dimensions[col].width = w

    return wb, ws


def _escrever_linha_regiao(ws, row: int, ano, regiao, total, pct, acomp, invalido=False, faltando=False):
    border = _thin_border()
    if faltando:
        fill = _FILL_MISSING
    elif invalido:
        fill = _FILL_INVALID
    else:
        fill = _FILL_REGIAO.get(str(regiao).upper(), PatternFill())

    for col, val in enumerate([ano, regiao, total, pct, acomp], 1):
        c = ws.cell(row=row, column=col, value=val)
        c.alignment = _center(); c.border = border; c.fill = fill


def processar_por_regiao():
    log.info("")
    log.info("=" * 60)
    log.info("PARTE 3: POR REGIÃO")
    log.info("=" * 60)

    source_dir = BASE_DIR / "por_regiao"
    if not source_dir.exists():
        log.error(f"Pasta não encontrada: {source_dir}")
        return

    existentes = _carregar_existentes_regiao(OUTPUT_REGIAO)
    log.info(f"Consolidado existente: {len(existentes)} registros")

    wb, ws = _criar_wb_regiao()
    # Anos disponíveis na pasta (pode ainda não ter todos os anos)
    arquivos = sorted(source_dir.glob("regiao_*.xls"))
    anos_disponiveis = ANOS  # tenta todos, marca faltando se não houver arquivo

    total_arq = len(anos_disponiveis)
    cont = {"ok": 0, "div": 0, "miss": 0, "skip": 0}
    divergencias_report = []
    row_num = 3

    for idx, ano in enumerate(anos_disponiveis, 1):
        nome = f"regiao_{ano}.xls"
        fp = source_dir / nome

        if not fp.exists():
            log.warning(f"  [{idx}/{total_arq}] FALTANDO: {nome}")
            cont["miss"] += 1
            # Grava uma linha placeholder por região
            for regiao in REGIOES:
                _escrever_linha_regiao(ws, row_num, ano, regiao, "ARQUIVO FALTANDO", None, None, faltando=True)
                row_num += 1
            continue

        # Verificar se TODAS as regiões deste ano já estão no consolidado
        chaves_existentes = [(ano, r) for r in REGIOES if (ano, r) in existentes
                             and existentes[(ano, r)]["total"] not in (None, "INVÁLIDO", "ARQUIVO FALTANDO")]
        if len(chaves_existentes) == len(REGIOES):
            log.info(f"  [{idx}/{total_arq}] JÁ PROCESSADO: {nome}")
            cont["skip"] += 1
            for regiao in REGIOES:
                d = existentes[(ano, regiao)]
                _escrever_linha_regiao(ws, row_num, ano, regiao, d["total"], d["percentual"], d["acompanhados"])
                row_num += 1
            continue

        log.info(f"  [{idx}/{total_arq}] Validando: {nome}")
        r = validar_por_regiao(fp, ano)

        if r["valido"]:
            cont["ok"] += 1
            for regiao in REGIOES:
                dados = r["regioes"].get(regiao, (None, None, None))
                log.info(f"    ✓ {regiao}: Total={dados[0]} | %={dados[1]} | Acomp={dados[2]}")
                _escrever_linha_regiao(ws, row_num, ano, regiao, dados[0], dados[1], dados[2])
                row_num += 1
        else:
            cont["div"] += 1
            log.warning(f"    ✗ DIVERGÊNCIAS em {nome}:")
            for d in r["divergencias"]:
                log.warning(f"      - {d}")
            divergencias_report.append((nome, r["divergencias"]))
            # Mesmo com divergências, gravar o que conseguiu extrair (ou INVÁLIDO)
            for regiao in REGIOES:
                dados = r["regioes"].get(regiao)
                if dados:
                    # Dados extraídos mesmo com divergência de metadados → grava mas marca vermelho
                    _escrever_linha_regiao(ws, row_num, ano, regiao, dados[0], dados[1], dados[2], invalido=True)
                else:
                    _escrever_linha_regiao(ws, row_num, ano, regiao, "INVÁLIDO", None, None, invalido=True)
                row_num += 1

    OUTPUT_REGIAO.parent.mkdir(parents=True, exist_ok=True)
    wb.save(OUTPUT_REGIAO)

    log.info(f"\n  Resultado: {cont['ok']} validados | {cont['skip']} reutilizados | "
             f"{cont['div']} divergências | {cont['miss']} faltando")
    log.info(f"  Salvo em: {OUTPUT_REGIAO.resolve()}")

    if divergencias_report:
        log.info("\n  ARQUIVOS COM DIVERGÊNCIAS (regiao):")
        for nome, divs in divergencias_report:
            log.warning(f"    {nome}:")
            for d in divs:
                log.warning(f"      - {d}")

    return cont


# ---------------------------------------------------------------------------
# Entry point
# ---------------------------------------------------------------------------

def main():
    parser = argparse.ArgumentParser(
        description="Validação e consolidação SISVAN (por_sexo, por_raca_cor, por_regiao)"
    )
    parser.add_argument("--sexo",   action="store_true", help="Processar apenas por_sexo")
    parser.add_argument("--raca",   action="store_true", help="Processar apenas por_raca_cor")
    parser.add_argument("--regiao", action="store_true", help="Processar apenas por_regiao")
    parser.add_argument("--todos",  action="store_true", help="Processar tudo (padrão)")
    args = parser.parse_args()

    # Por padrão (sem flags), processa tudo
    alguma_flag = args.sexo or args.raca or args.regiao
    rodar_sexo   = args.todos or args.sexo   or not alguma_flag
    rodar_raca   = args.todos or args.raca   or not alguma_flag
    rodar_regiao = args.todos or args.regiao or not alguma_flag

    log.info("SISVAN - Validação e Consolidação de Dados")
    log.info(f"Anos: {ANOS[0]}–{ANOS[-1]}")

    if rodar_sexo:
        processar_por_sexo()
    if rodar_raca:
        processar_por_raca()
    if rodar_regiao:
        processar_por_regiao()

    log.info("")
    log.info("=== CONCLUÍDO ===")


if __name__ == "__main__":
    main()
